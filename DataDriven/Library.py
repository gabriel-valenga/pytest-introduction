import json
import jsonpath
from openpyxl import load_workbook

class ExcelSheet:

    def __init__(self, fileNamePath, SheetName):
        global workbook
        global sheet
        workbook = load_workbook(fileNamePath)
        sheet = workbook[SheetName]


    def fetch_sheet_row_count(self):
        return sheet.max_row
    

    def fetch_sheet_column_count(self):
        return sheet.max_column


    def fetch_sheet_key_names(self):
        sheet_max_columns = self.fetch_sheet_column_count()
        key_names= []
        for i in range(i, sheet_max_columns + 1):
            cell = sheet.cell(row=1, column=i)
            key_names.insert(cell.value)
        return key_names
    

    def format_json_request(self, row_number):
        json_request = {}
        sheet_max_columns = self.fetch_sheet_column_count()
        key_names = self.fetch_sheet_key_names()
        for i in range(1, sheet_max_columns + 1):
            cell = sheet.cell(row=row_number, column=i)
            json_request.update({key_names[i-1]: cell.value})
